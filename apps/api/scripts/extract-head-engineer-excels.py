#!/usr/bin/env python3
"""Extract head-engineer Excel trackers into JSON seed files."""
from __future__ import annotations

import json
import re
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[3]
SAMPLE_DIR = REPO_ROOT / "sample study"
OUT_CONCRETE = REPO_ROOT / "apps" / "api" / "prisma" / "seed-data" / "jamila-concrete-rebar.json"
OUT_SCHEDULE = REPO_ROOT / "apps" / "api" / "prisma" / "seed-data" / "jura-schedule.json"

REBAR_DIAMETERS = ["6", "8", "10", "12", "16", "18", "22", "25"]
EXCEL_EPOCH = datetime(1899, 12, 30)


def excel_date(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str):
        s = value.strip()
        if not s or s in ("Done", "Done ", "جاري", "جاري ", "___"):
            return None
        m = re.match(r"^(\d{1,2})-(\d{1,2})-(\d{4})$", s)
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return f"{y:04d}-{mo:02d}-{d:02d}"
        return None
    if isinstance(value, (int, float)) and value > 40000:
        return (EXCEL_EPOCH + timedelta(days=float(value))).date().isoformat()
    return None


def infer_status(*parts) -> str:
    text = " ".join(str(p) for p in parts if p is not None).lower()
    if "جاري" in text or "in progress" in text:
        return "IN_PROGRESS"
    if "done" in text:
        return "DONE"
    return "PLANNED"


def num(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    if not s or s in ("Done", "Done ", "___", "جاري", "جاري "):
        return None
    try:
        return float(s)
    except ValueError:
        return None


REBAR_COLS = {"6": 4, "8": 5, "10": 6, "12": 7, "16": 8, "18": 9, "22": 10, "25": 11}


def extract_concrete_rebar(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    entries: list[dict] = []
    building = ""
    half_zone = ""
    floor = ""

    for idx, row in enumerate(rows, start=1):
        if idx < 4:
            continue
        cells = list(row) + [None] * 25

        if cells[0] and str(cells[0]).strip():
            building = str(cells[0]).strip()
        if cells[1] and str(cells[1]).strip() and "النصف" in str(cells[1]):
            half_zone = str(cells[1]).strip()
        if cells[2] and str(cells[2]).strip():
            floor = str(cells[2]).strip()

        element = cells[3]
        if not element or not str(element).strip():
            continue
        element_s = str(element).strip()

        if not building:
            continue

        rebar: dict[str, float] = {}
        for dia, col in REBAR_COLS.items():
            val = num(cells[col])
            if val and val > 0:
                rebar[dia] = val

        concrete = num(cells[14])
        rebar_cost = num(cells[15])
        concrete_cost = num(cells[16])
        duration = num(cells[17])
        start = excel_date(cells[18])
        end = excel_date(cells[19])
        actual = excel_date(cells[20]) or excel_date(cells[21])

        if element_s.startswith("سقف") or "سقف" in element_s and "اعمدة" not in element_s:
            element_type = "SLAB"
        elif "اعمدة" in element_s and "سقف" in element_s:
            element_type = "COMBINED"
        elif "اعمدة" in element_s:
            element_type = "COLUMNS"
        elif "قواعد" in element_s or "سملات" in element_s:
            element_type = "FOUNDATION"
        elif "مباني" in element_s or "تشطيب" in element_s:
            element_type = "OTHER"
        else:
            element_type = "OTHER"

        status = "IN_PROGRESS" if "جاري" in element_s else infer_status(cells[4], actual)
        if actual:
            status = "DONE"
        elif isinstance(cells[4], str) and "جاري" in str(cells[4]):
            status = "IN_PROGRESS"

        if not rebar and not concrete and not rebar_cost and not concrete_cost:
            continue

        entries.append(
            {
                "buildingLabel": building,
                "halfZone": half_zone or None,
                "floorLevel": floor or None,
                "elementType": element_type,
                "elementLabel": element_s,
                "rebarByDiameter": rebar,
                "concreteM3": concrete,
                "rebarCostEGP": rebar_cost,
                "concreteCostEGP": concrete_cost,
                "laborCostEGP": None,
                "plannedDurationDays": int(duration) if duration else None,
                "plannedStart": start,
                "plannedEnd": end,
                "actualPourDate": actual,
                "status": status,
                "sourceRow": idx,
            }
        )

    return entries


def extract_jura_schedule(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    buildings = ["B05", "B06", "B07"]
    building_cols = {"B05": 4, "B06": 9, "B07": 14}
    plan_name = path.stem
    deadline = None
    lines: list[dict] = []
    sort_order = 0
    current_category = ""

    for idx, row in enumerate(rows, start=1):
        cells = list(row) + [None] * 30
        if idx == 6:
            for v in cells:
                d = excel_date(v)
                if d:
                    deadline = d
        if idx < 11:
            continue

        item_code = cells[1]
        if item_code is None:
            continue

        if isinstance(item_code, (int, float)):
            code_s = str(int(item_code)) if float(item_code) == int(float(item_code)) else str(item_code)
        else:
            code_s = str(item_code).strip()

        if re.match(r"^\d+$", code_s) and not cells[3]:
            current_category = str(cells[2] or "").strip()
            continue

        if "-" not in code_s:
            continue

        desc = str(cells[2] or "").strip()
        unit_raw = cells[3]
        unit = str(unit_raw).strip() if unit_raw else None
        if unit in ("None", ""):
            unit = None

        progress = []
        for b in buildings:
            base = building_cols[b]
            qty = num(cells[base])
            rate = num(cells[base + 1])
            days = num(cells[base + 2])
            start = excel_date(cells[base + 3])
            end = excel_date(cells[base + 4])
            status_raw = cells[base]
            status = None
            if isinstance(status_raw, str) and "done" in status_raw.lower():
                status = "DONE"
            elif start and end and qty:
                status = "PLANNED"
            progress.append(
                {
                    "buildingCode": b,
                    "quantity": qty,
                    "rateEGP": rate,
                    "durationDays": int(days) if days else None,
                    "startDate": start,
                    "endDate": end,
                    "lineTotalEGP": (qty or 0) * (rate or 0) if qty and rate else None,
                    "status": status or infer_status(status_raw, start, end),
                }
            )

        parent = code_s.split("-")[0] if "-" in code_s else None
        sort_order += 1
        lines.append(
            {
                "itemCode": code_s,
                "parentCode": parent,
                "description": desc,
                "unit": unit,
                "categoryLabel": current_category or None,
                "sortOrder": sort_order,
                "progress": progress,
            }
        )

    return {
        "plan": {
            "name": plan_name,
            "projectLabel": "Jura Phase 1 — Buildings 5, 6, 7",
            "planDeadline": deadline,
            "buildingCodes": buildings,
        },
        "lines": lines,
    }


def main() -> None:
    concrete_path = SAMPLE_DIR / "Jamila Concrete and Reinforcement 2.xlsx"
    jura_path = SAMPLE_DIR / "Jura PH1 Plan till 30-Nov-2026 - H.xlsx"

    if not concrete_path.exists():
        raise SystemExit(f"Missing: {concrete_path}")
    if not jura_path.exists():
        raise SystemExit(f"Missing: {jura_path}")

    concrete = extract_concrete_rebar(concrete_path)
    schedule = extract_jura_schedule(jura_path)

    OUT_CONCRETE.parent.mkdir(parents=True, exist_ok=True)
    OUT_CONCRETE.write_text(json.dumps(concrete, ensure_ascii=False, indent=2), encoding="utf-8")
    OUT_SCHEDULE.write_text(json.dumps(schedule, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Wrote {len(concrete)} pour tracker entries -> {OUT_CONCRETE.name}")
    print(f"Wrote {len(schedule['lines'])} schedule lines -> {OUT_SCHEDULE.name}")


if __name__ == "__main__":
    main()
