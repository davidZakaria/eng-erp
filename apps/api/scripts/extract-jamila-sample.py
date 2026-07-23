#!/usr/bin/env python3
"""
Regenerate apps/api/prisma/seed-data/jamila-from-sample.json from local
sample study/excel files (not committed to git).

Usage (from repo root):
  py -3 apps/api/scripts/extract-jamila-sample.py
"""
from __future__ import annotations

import json
import os
import re
from pathlib import Path

from openpyxl import load_workbook
import xlrd

REPO_ROOT = Path(__file__).resolve().parents[3]
SAMPLE_DIR = REPO_ROOT / "sample study" / "excel"
OUT_PATH = REPO_ROOT / "apps" / "api" / "prisma" / "seed-data" / "jamila-from-sample.json"


def normalize_unit(unit: str) -> str:
    return (
        str(unit)
        .replace("\ufffd", "²")
        .replace("m2", "m²")
        .replace("m3", "m³")
        .strip()
    )


def parse_spec_filename(name: str) -> dict | None:
    base = Path(name).stem
    m = re.match(r"^SECTION\s+(\d{5})[-\s]+(.+)$", base, re.I)
    if m:
        code = m.group(1)
        return {
            "code": code,
            "title": m.group(2).strip(),
            "divisionCode": code[:2],
        }
    m = re.match(r"^(\d{6,8})\s+(.+)$", base)
    if m:
        code = m.group(1)
        return {
            "code": code,
            "title": m.group(2).strip(),
            "divisionCode": code[:2],
        }
    return None


def extract_boq_rows(ws, division_code: str | None = None) -> list[dict]:
    items: list[dict] = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        vals = list(row) + [None] * 6
        item, desc, unit, qty, rate = vals[:5]
        if item is None or desc is None:
            continue
        item_s = str(item).strip()
        if not re.match(r"^\d+(\.\d+)+$", item_s):
            continue
        if not unit or not isinstance(qty, (int, float)) or qty <= 0:
            continue
        div = division_code
        if not div and len(vals) > 6:
            div = None
        items.append(
            {
                "itemCode": item_s,
                "description": str(desc).strip()[:240],
                "unit": normalize_unit(str(unit)),
                "plannedQuantity": float(qty),
                "rateEGP": float(rate) if isinstance(rate, (int, float)) else 0,
                "divisionCode": div or item_s.split(".")[0].zfill(2),
            }
        )
    return items


def main() -> None:
    if not SAMPLE_DIR.is_dir():
        raise SystemExit(f"Sample folder not found: {SAMPLE_DIR}")

    out: dict = {
        "project": {
            "name": "Jamila Residences",
            "code": "JAM-001",
            "client": "New Jersey Developments",
            "location": "North Coast, Egypt",
            "buildingType": "G1",
            "projectNumber": "2315",
        },
        "csiDivisions": [
            {"code": "02", "title": "Site Construction / Earthwork"},
            {"code": "03", "title": "Concrete"},
            {"code": "04", "title": "Masonry"},
            {"code": "05", "title": "Metals"},
            {"code": "07", "title": "Thermal and Moisture Protection"},
            {"code": "08", "title": "Openings"},
            {"code": "09", "title": "Finishes"},
            {"code": "10", "title": "Specialties"},
            {"code": "14", "title": "Conveying Equipment"},
            {"code": "21", "title": "Fire Suppression"},
            {"code": "22", "title": "Plumbing"},
            {"code": "23", "title": "HVAC"},
            {"code": "26", "title": "Electrical"},
        ],
        "vendors": [
            {"name": "Egypt Ready Mix Co.", "country": "Egypt", "divisionCode": "03", "disciplineTag": "Concrete"},
            {"name": "Cairo Climate Systems", "country": "Egypt", "divisionCode": "23", "disciplineTag": "HVAC"},
            {"name": "Delta Fire Systems", "country": "Egypt", "divisionCode": "21", "disciplineTag": "Fire Fighting"},
            {"name": "Nile Plumbing Supplies", "country": "Egypt", "divisionCode": "22", "disciplineTag": "Plumbing"},
            {"name": "Alexandria Steel Works", "country": "Egypt", "divisionCode": "05", "disciplineTag": "Structural Steel"},
        ],
    }

    # Architectural LOD
    wb = load_workbook(SAMPLE_DIR / "2315-A-002-LOD-G1.xlsx", read_only=True, data_only=True)
    ws = wb.active
    arch_lod = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        vals = list(row)
        prj, disc, sheet, title, rev, size, scale, status = (vals + [None] * 8)[1:9]
        if disc and sheet and title and str(disc).strip() not in ("DIS.\nDES.#",):
            if isinstance(disc, str) and disc.strip().startswith("0."):
                continue
            arch_lod.append(
                {
                    "projectNo": str(prj) if prj else "2315",
                    "disciplineCode": str(disc).strip(),
                    "sheetNumber": str(sheet).strip(),
                    "title": str(title).strip(),
                    "revision": int(rev) if isinstance(rev, (int, float)) else 0,
                    "size": str(size).strip() if size else None,
                    "scale": str(scale).strip() if scale else None,
                    "status": str(status).strip() if status else None,
                }
            )
    wb.close()
    out["architecturalLod"] = arch_lod

    # Electrical LOD
    wb = load_workbook(SAMPLE_DIR / "2018-E-002-LOD-G1.xlsx", read_only=True, data_only=True)
    ws = wb["L.O.D"]
    elec_lod = []
    for row in ws.iter_rows(min_row=11, values_only=True):
        vals = list(row)
        prj, disc, sheet, title, rev, size, scale = (vals + [None] * 8)[1:8]
        if not disc or not sheet or not title:
            continue
        disc_s = str(disc).strip()
        if "DETAIL" in disc_s.upper():
            continue
        elec_lod.append(
            {
                "projectNo": str(prj) if prj else "2315",
                "disciplineCode": disc_s,
                "sheetNumber": str(sheet).strip(),
                "title": str(title).strip(),
                "revision": int(rev) if isinstance(rev, (int, float)) else 0,
                "size": str(size).strip() if size else None,
                "scale": str(scale).strip() if scale else None,
                "status": None,
            }
        )
    wb.close()
    out["electricalLod"] = elec_lod

    # Architectural BOQ
    wb = load_workbook(SAMPLE_DIR / "BOQ -type G1-REV 00.xlsx", read_only=True, data_only=True)
    arch_boq: list[dict] = []
    for sheet in wb.sheetnames:
        if not sheet.startswith("Div"):
            continue
        div_code = re.sub(r"\D", "", sheet)[:2] or sheet.replace("Div.", "")
        arch_boq.extend(extract_boq_rows(wb[sheet], div_code))
    wb.close()
    out["architecturalBoq"] = arch_boq

    # Structural BOQ
    wb = load_workbook(SAMPLE_DIR / "G1- BOQ 2 Structural 02.xlsx", read_only=True, data_only=True)
    out["structuralBoq"] = extract_boq_rows(wb["Bill of Quantities"], "03")
    wb.close()

    # Plumbing BOQ
    wb = load_workbook(SAMPLE_DIR / "PLUMBING  BOQ. G1.xlsx", read_only=True, data_only=True)
    plumb = extract_boq_rows(wb["Division 22"], "22")
    for item in plumb:
        item["itemCode"] = f"22-{item['itemCode']}"
    out["plumbingBoq"] = plumb
    wb.close()

    # Fire fighting BOQ
    wb = load_workbook(SAMPLE_DIR / "BOQ-FF G1.xlsx", read_only=True, data_only=True)
    ff: list[dict] = []
    for sheet in wb.sheetnames:
        if sheet.lower().startswith("cover"):
            continue
        for item in extract_boq_rows(wb[sheet], "21"):
            item["itemCode"] = f"FF-{item['itemCode']}"
            ff.append(item)
    wb.close()
    out["fireFightingBoq"] = ff

    # Load schedules
    wb = xlrd.open_workbook(str(SAMPLE_DIR / "LOAD SCHEDULES (G1).xls"))
    panels = []
    for name in wb.sheet_names():
        sh = wb.sheet_by_name(name)
        ref = str(sh.cell_value(0, 4) or name).strip()
        loc = str(sh.cell_value(1, 4)).strip()
        incoming = str(sh.cell_value(2, 4)).strip() or "40A"
        circuits = []
        for r in range(6, sh.nrows):
            cb = sh.cell_value(r, 0)
            if not isinstance(cb, (int, float)) or cb == 0:
                continue
            mcb = sh.cell_value(r, 1)
            wire = sh.cell_value(r, 2)
            load_type = sh.cell_value(r, 3)
            load_va = sh.cell_value(r, 4)
            df = sh.cell_value(r, 8)
            rva, yva, bva = sh.cell_value(r, 5), sh.cell_value(r, 6), sh.cell_value(r, 7)
            phase = "R"
            if isinstance(yva, (int, float)) and yva:
                phase = "Y"
            elif isinstance(bva, (int, float)) and bva:
                phase = "B"
            try:
                mcb_f = float(str(mcb).replace("#", ""))
            except ValueError:
                mcb_f = 16.0
            circuits.append(
                {
                    "circuitNumber": int(cb),
                    "mcbRating": mcb_f,
                    "wireSize": str(wire),
                    "loadType": str(load_type) if load_type else "General",
                    "connectedLoadVA": float(load_va) if isinstance(load_va, (int, float)) else 0,
                    "demandFactor": float(df) if isinstance(df, (int, float)) and df else 0.8,
                    "phase": phase,
                }
            )
        panels.append(
            {
                "panelReference": ref,
                "location": f"Building G1 {loc}",
                "incomingCB": incoming,
                "circuits": circuits,
            }
        )
    out["electricalPanels"] = panels

    # Spec sections from PDF filenames
    spec_sections: dict[str, dict] = {}
    for path in SAMPLE_DIR.glob("*.pdf"):
        parsed = parse_spec_filename(path.name)
        if parsed and parsed["code"] not in spec_sections:
            spec_sections[parsed["code"]] = parsed
    out["specSections"] = sorted(spec_sections.values(), key=lambda s: s["code"])

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUT_PATH.open("w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"Wrote {OUT_PATH}")
    print(
        "Counts:",
        {k: len(v) if isinstance(v, list) else "object" for k, v in out.items()},
    )


if __name__ == "__main__":
    main()
